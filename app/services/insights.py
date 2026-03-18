"""Learning & insights service.

Queries the extraction_insights table for prompt hints, processes feedback
to learn from corrections, and reads CSV/XLSX files from screenshot directories.
"""

import csv
import json
import logging
from pathlib import Path

from app.config import settings
from app.database import get_db

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Calibration data — baked-in from 16 DD + 11 NR real SE spreadsheets
# These are injected into GPT-4o Vision prompts for better accuracy.
# ---------------------------------------------------------------------------
DD_CALIBRATION = {
    "customer_count": 16,
    "field_ranges": {
        "infra_hosts": {"min": 0, "max": 4094, "typical": "200–1500", "unit": "count or hourly"},
        "apm_hosts": {"min": 0, "max": 700, "typical": "20–300", "unit": "count or hourly"},
        "containers": {"min": 0, "max": 15431, "typical": "1800–5500", "unit": "hourly"},
        "custom_metrics": {"min": 135, "max": 435417, "typical": "2000–200000", "unit": "count or hourly"},
        "ingested_logs_gb": {"min": 0, "max": 404000, "typical": "1000–52000", "unit": "GB/month"},
        "ingested_spans_gb": {"min": 0, "max": 67000, "typical": "500–16000", "unit": "GB/month (check TB suffix!)"},
        "indexed_spans_million": {"min": 0, "max": 1740, "typical": "10–200", "unit": "millions/month"},
        "rum_sessions": {"min": 0, "max": 5296, "typical": "0–5000", "unit": "count/month"},
        "serverless_invocations": {"min": 0, "max": 21348684, "typical": "0–250000", "unit": "count/month"},
    },
    "constants": {
        "AVG_LOG_SIZE_KB": "2.5 (standard), 2.0 (conservative), 1.0 (minimal)",
        "AVG_SPAN_SIZE_KB": "1.5 (standard), 2.0–2.5 (verbose)",
        "TS_PER_NODE": 750,
        "hourly_divisor": "~655 (30.4 × 24 × 0.9) or /30/24×1.1",
    },
    "tier_distributions": {
        "default (Active)": "FS 50% / Mon 40% / Comp 10%",
        "BigId (Conservative)": "FS 10% / Mon 70% / Comp 20%",
    },
}

NR_CALIBRATION = {
    "customer_count": 11,
    "field_ranges": {
        "logging_gb_day": {"min": 0, "max": 4187, "typical": "2–500"},
        "metrics_gb_day": {"min": 9, "max": 7433, "typical": "40–1000"},
        "infra_integrations_gb_day": {"min": 0, "max": 2300, "typical": "13–450"},
        "infra_hosts_gb_day": {"min": 0, "max": 95, "typical": "1–50"},
        "infra_processes_gb_day": {"min": 0, "max": 1170, "typical": "0–300"},
        "apm_events_gb_day": {"min": 0, "max": 7753, "typical": "7–1700"},
        "tracing_gb_day": {"min": 0, "max": 3942, "typical": "10–330"},
        "browser_events_gb_day": {"min": 0, "max": 504, "typical": "0–40"},
        "custom_events_gb_day": {"min": 0, "max": 3145, "typical": "0–50"},
    },
    "constants": {
        "NumSeries": "total_metrics_gb × 1000",
        "RUM_sessions_from_GB": "(GB × 1,000,000) / 800",
    },
}


CW_CALIBRATION = {
    "customer_count": 0,  # No calibration data yet — new provider
    "field_ranges": {
        "total_put_log_events_gb": {"min": 0, "max": 50000, "typical": "50–5000", "unit": "GB/month"},
        "total_custom_metrics_count": {"min": 0, "max": 500000, "typical": "1000–50000", "unit": "count"},
        "total_alarms_count": {"min": 0, "max": 10000, "typical": "10–500", "unit": "count"},
        "total_xray_traces": {"min": 0, "max": 100000000, "typical": "0–10000000", "unit": "count/month"},
        "total_xray_segments": {"min": 0, "max": 500000000, "typical": "0–50000000", "unit": "count/month"},
        "total_start_query_gb": {"min": 0, "max": 10000, "typical": "0–500", "unit": "GB/month"},
    },
    "constants": {
        "CX_LOG_ENRICHMENT_MULTIPLIER": "1.77 (455 bytes CW → 805 bytes CX)",
        "XRAY_KB_PER_TRACE": 5,
        "XRAY_KB_PER_SEGMENT": 2,
    },
}


def get_hints(provider: str) -> list[str]:
    """Return accumulated correction hints + calibration data for extraction prompts."""
    # Start with baked-in calibration from real spreadsheets
    hints = get_calibration_hints(provider)

    # Add DB-learned hints from past feedback
    with get_db() as db:
        rows = db.execute(
            "SELECT correction_hint, sample_count FROM extraction_insights "
            "WHERE provider = ? AND field_name NOT LIKE '/_%%' ESCAPE '/' "
            "ORDER BY sample_count DESC LIMIT 15",
            (provider,),
        ).fetchall()
    hints.extend(row["correction_hint"] for row in rows)
    return hints


def record_feedback(
    run_id: int,
    is_accurate: bool,
    notes: str | None,
    field_corrections: dict | None,
) -> None:
    """Store user feedback and update extraction_insights from corrections."""
    with get_db() as db:
        db.execute(
            "INSERT INTO feedback (run_id, is_accurate, notes, field_corrections) "
            "VALUES (?, ?, ?, ?)",
            (run_id, is_accurate, notes, json.dumps(field_corrections) if field_corrections else None),
        )

        # Update the run status
        db.execute(
            "UPDATE sizing_runs SET status = 'feedback_received' WHERE id = ?",
            (run_id,),
        )

        if not field_corrections:
            return

        # Get the provider for this run
        row = db.execute("SELECT provider FROM sizing_runs WHERE id = ?", (run_id,)).fetchone()
        if not row:
            return
        provider = row["provider"]

        # Upsert insights from each field correction
        for field_name, correction in field_corrections.items():
            extracted = correction.get("extracted")
            actual = correction.get("actual")
            if extracted is None or actual is None:
                continue

            common_error = f"Extracted {extracted} but actual was {actual}"
            correction_hint = (
                f"For '{field_name}': previously extracted as {extracted} "
                f"but the correct value was {actual}. Double-check this field."
            )

            db.execute(
                """INSERT INTO extraction_insights
                   (provider, field_name, common_error, correction_hint, sample_count)
                   VALUES (?, ?, ?, ?, 1)
                   ON CONFLICT(provider, field_name, common_error) DO UPDATE SET
                     sample_count = sample_count + 1,
                     last_updated = CURRENT_TIMESTAMP""",
                (provider, field_name, common_error, correction_hint),
            )


def get_all_insights(provider: str | None = None) -> list[dict]:
    """Return all extraction insights, optionally filtered by provider."""
    with get_db() as db:
        if provider:
            rows = db.execute(
                "SELECT * FROM extraction_insights WHERE provider = ? "
                "ORDER BY sample_count DESC, last_updated DESC",
                (provider,),
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM extraction_insights "
                "ORDER BY sample_count DESC, last_updated DESC"
            ).fetchall()
    return [dict(row) for row in rows]


def get_calibration_hints(provider: str) -> list[str]:
    """Return baked-in calibration hints for a provider's extraction prompt."""
    if provider == "datadog":
        cal = DD_CALIBRATION
    elif provider == "cloudwatch":
        cal = CW_CALIBRATION
    else:
        cal = NR_CALIBRATION
    hints = [
        f"Based on {cal['customer_count']} real customer sizing spreadsheets:"
    ]
    for field, info in cal["field_ranges"].items():
        typical = info.get("typical", "")
        unit = info.get("unit", "")
        suffix = f" ({unit})" if unit else ""
        hints.append(f"  {field}: typical range {typical}{suffix}")
    return hints


def learn_from_data_files() -> dict[str, int]:
    """Scan screenshot directories for CSV and XLSX files and learn field patterns.

    Returns a dict like {"datadog": 3, "newrelic": 1} with count of files processed.
    """
    results: dict[str, int] = {}

    for provider in ("datadog", "newrelic", "cloudwatch"):
        provider_dir = settings.screenshots_dir / provider
        if not provider_dir.exists():
            continue

        count = 0

        # Process CSV files
        for csv_path in provider_dir.glob("*.csv"):
            try:
                _process_csv(provider, csv_path)
                count += 1
            except Exception as e:
                logger.warning("Failed to process CSV %s: %s", csv_path, e)

        # Process XLSX files
        for xlsx_path in provider_dir.glob("*.xlsx"):
            try:
                _process_xlsx(provider, xlsx_path)
                count += 1
            except Exception as e:
                logger.warning("Failed to process XLSX %s: %s", xlsx_path, e)

        results[provider] = count

    return results


# Keep old name as alias for backward compat
learn_from_csv_files = learn_from_data_files


def _process_csv(provider: str, csv_path: Path) -> None:
    """Extract field-value patterns from a sizing CSV to build insights."""
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) < 2:
        return

    # Try to find header row and value rows
    # Common CSV formats: either key-value pairs or header + data rows
    header = rows[0]

    with get_db() as db:
        # Store a hint that this CSV was processed
        db.execute(
            """INSERT INTO extraction_insights
               (provider, field_name, common_error, correction_hint, sample_count)
               VALUES (?, ?, ?, ?, 1)
               ON CONFLICT(provider, field_name, common_error) DO UPDATE SET
                 sample_count = sample_count + 1,
                 last_updated = CURRENT_TIMESTAMP""",
            (
                provider,
                "_csv_source",
                f"Learned from {csv_path.name}",
                f"Reference CSV '{csv_path.name}' available with {len(rows)-1} data rows. "
                f"Columns: {', '.join(header[:10])}",
            ),
        )

        # Extract specific field patterns from known CSV formats
        _extract_csv_patterns(db, provider, header, rows[1:])


def _extract_csv_patterns(
    db, provider: str, header: list[str], data_rows: list[list[str]]
) -> None:
    """Extract reusable patterns from CSV data rows."""
    # Normalize headers
    norm_header = [h.strip().lower() for h in header]

    # Look for key columns that map to our extraction fields
    field_mappings_dd = {
        "infra host": "infra_hosts",
        "apm host": "apm_hosts",
        "container": "containers",
        "custom metric": "custom_metrics",
        "ingested log": "ingested_logs_gb",
        "indexed log": "indexed_logs",
        "ingested span": "ingested_spans_gb",
        "indexed span": "indexed_spans_million",
        "rum session": "rum_sessions",
        "total metric": "total_metrics_from_overview",
    }

    field_mappings_nr = {
        "logging": "logging_gb_day",
        "custom event": "custom_events_gb_day",
        "metric": "metrics_gb_day",
        "infra": "infra_hosts_gb_day",
        "apm": "apm_events_gb_day",
        "tracing": "tracing_gb_day",
        "browser": "browser_events_gb_day",
    }

    mappings = field_mappings_dd if provider == "datadog" else field_mappings_nr

    for col_idx, col_name in enumerate(norm_header):
        for pattern, field_name in mappings.items():
            if pattern in col_name:
                # Extract non-empty values from this column
                values = []
                for row in data_rows:
                    if col_idx < len(row) and row[col_idx].strip():
                        try:
                            val = row[col_idx].strip().replace(",", "")
                            float(val)
                            values.append(val)
                        except ValueError:
                            continue

                if values:
                    hint = (
                        f"CSV reference shows '{col_name}' values like: "
                        f"{', '.join(values[:3])}. "
                        f"Typical range for {field_name}."
                    )
                    db.execute(
                        """INSERT INTO extraction_insights
                           (provider, field_name, common_error, correction_hint, sample_count)
                           VALUES (?, ?, ?, ?, 1)
                           ON CONFLICT(provider, field_name, common_error) DO UPDATE SET
                             sample_count = sample_count + 1,
                             last_updated = CURRENT_TIMESTAMP""",
                        (provider, field_name, f"CSV range from {col_name}", hint),
                    )
                break


def _process_xlsx(provider: str, xlsx_path: Path) -> None:
    """Extract field-value patterns from a sizing XLSX to build insights."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — skipping XLSX learning")
        return

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)

    with get_db() as db:
        # Record that we processed this file
        db.execute(
            """INSERT INTO extraction_insights
               (provider, field_name, common_error, correction_hint, sample_count)
               VALUES (?, ?, ?, ?, 1)
               ON CONFLICT(provider, field_name, common_error) DO UPDATE SET
                 sample_count = sample_count + 1,
                 last_updated = CURRENT_TIMESTAMP""",
            (
                provider,
                "_xlsx_source",
                f"Learned from {xlsx_path.name}",
                f"Reference XLSX '{xlsx_path.name}' with sheets: "
                f"{', '.join(wb.sheetnames[:5])}",
            ),
        )

        # Try to find the conversion sheet (DD) or sizing data (NR)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            if len(rows) < 2:
                continue

            # Scan for known field names and extract adjacent values
            if provider == "datadog":
                _learn_dd_xlsx_values(db, rows, xlsx_path.name)
            else:
                _learn_nr_xlsx_values(db, rows, xlsx_path.name)

    wb.close()


def _learn_dd_xlsx_values(db, rows: list[list[str]], filename: str) -> None:
    """Learn Datadog field values from an XLSX sheet."""
    dd_patterns = {
        "infra host": "infra_hosts",
        "apm host": "apm_hosts",
        "container": "containers",
        "custom metric": "custom_metrics",
        "ingested log": "ingested_logs_gb",
        "ingested span": "ingested_spans_gb",
        "indexed span": "indexed_spans_million",
        "serverless inv": "serverless_invocations",
        "serverless func": "serverless_functions",
        "rum session": "rum_sessions",
        "logs gb": "logs_gb_day",
        "tracing gb": "traces_gb_day",
        "numseries": "metrics_num_series",
    }

    for row in rows:
        if not row:
            continue
        label = row[0].strip().lower()
        for pattern, field_name in dd_patterns.items():
            if pattern in label:
                # Look for a numeric value in adjacent cells
                for cell in row[1:5]:
                    cell_val = cell.strip().replace(",", "").replace("$", "")
                    try:
                        val = float(cell_val)
                        if val > 0:
                            hint = (
                                f"Customer '{filename}': {field_name} = {val:,.0f}. "
                                f"Use as reference for range validation."
                            )
                            db.execute(
                                """INSERT INTO extraction_insights
                                   (provider, field_name, common_error, correction_hint, sample_count)
                                   VALUES (?, ?, ?, ?, 1)
                                   ON CONFLICT(provider, field_name, common_error) DO UPDATE SET
                                     sample_count = sample_count + 1,
                                     last_updated = CURRENT_TIMESTAMP""",
                                ("datadog", field_name, f"XLSX ref {filename}: {label}", hint),
                            )
                            break
                    except (ValueError, TypeError):
                        continue
                break


def _learn_nr_xlsx_values(db, rows: list[list[str]], filename: str) -> None:
    """Learn New Relic field values from an XLSX sheet."""
    nr_patterns = {
        "logging": "logging_gb_day",
        "custom event": "custom_events_gb_day",
        "serverless": "serverless_gb_day",
        "metric": "metrics_gb_day",
        "infrastructure integration": "infra_integrations_gb_day",
        "infra integration": "infra_integrations_gb_day",
        "infrastructure host": "infra_hosts_gb_day",
        "infra host": "infra_hosts_gb_day",
        "infrastructure process": "infra_processes_gb_day",
        "infra process": "infra_processes_gb_day",
        "apm event": "apm_events_gb_day",
        "tracing": "tracing_gb_day",
        "browser event": "browser_events_gb_day",
        "mobile event": "mobile_events_gb_day",
    }

    for row in rows:
        if not row:
            continue
        label = row[0].strip().lower()
        for pattern, field_name in nr_patterns.items():
            if pattern in label:
                for cell in row[1:5]:
                    cell_val = cell.strip().replace(",", "").replace("$", "")
                    try:
                        val = float(cell_val)
                        if val > 0:
                            hint = (
                                f"Customer '{filename}': {field_name} = {val:,.2f} GB/day. "
                                f"Use as reference."
                            )
                            db.execute(
                                """INSERT INTO extraction_insights
                                   (provider, field_name, common_error, correction_hint, sample_count)
                                   VALUES (?, ?, ?, ?, 1)
                                   ON CONFLICT(provider, field_name, common_error) DO UPDATE SET
                                     sample_count = sample_count + 1,
                                     last_updated = CURRENT_TIMESTAMP""",
                                ("newrelic", field_name, f"XLSX ref {filename}: {label}", hint),
                            )
                            break
                    except (ValueError, TypeError):
                        continue
                break


def get_run_history(limit: int = 50, user_email: str | None = None) -> list[dict]:
    """Return recent sizing runs with their feedback status, optionally filtered by user."""
    with get_db() as db:
        if user_email:
            rows = db.execute(
                """SELECT s.*,
                          f.is_accurate as feedback_accurate,
                          f.notes as feedback_notes
                   FROM sizing_runs s
                   LEFT JOIN feedback f ON f.run_id = s.id
                   WHERE s.user_email = ?
                   ORDER BY s.created_at DESC
                   LIMIT ?""",
                (user_email, limit),
            ).fetchall()
        else:
            rows = db.execute(
                """SELECT s.*,
                          f.is_accurate as feedback_accurate,
                          f.notes as feedback_notes
                   FROM sizing_runs s
                   LEFT JOIN feedback f ON f.run_id = s.id
                   ORDER BY s.created_at DESC
                   LIMIT ?""",
                (limit,),
            ).fetchall()
    return [dict(row) for row in rows]
